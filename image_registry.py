#Import libraries

import glob
import io,os
import time
from datetime import datetime
from datetime import timedelta
import requests
import numpy as np
import csv
import cv2
import os.path
from os import path
from zipfile import ZipFile
from requests.auth import HTTPBasicAuth
from bs4 import BeautifulSoup
import pandas as pd
from ibm_botocore.client import Config
import ibm_boto3
from pathlib import Path
import re
from requests import ConnectionError
from shutil import copyfile
import shutil
import requests
from PIL import Image
from os import listdir
from PIL import Image as PImage
import openpyxl
#from ibm_s3transfer.aspera.manager import AsperaTransferManager,AsperaConfig

cosCredentials = {
    'IBM_API_KEY_ID': 'QF2fRbhHUW_7nDOaS08TGyC_z8AWZ1mdi3YYdXmuHB4q',
    'IAM_SERVICE_ID': 'crn:v1:bluemix:public:iam-identity::a/c49a90cbc1224aaa935a27c8d35cec76::serviceid:ServiceId-e27d06dd-80d6-4244-9d2f-195e9358750c',
    'ENDPOINT': 'https://s3.us-south.cloud-object-storage.appdomain.cloud',
    'IBM_AUTH_ENDPOINT': 'https://iam.cloud.ibm.com/identity/token'

}

# s3.us.cloud-object-storage.appdomain.cloud
#Cos Bucket Name

defCOSBucketName = 'testannotateupload'

match_count=0
#features list
allfeatures=[11,15,28,33,32,49,39]
flist=[11,15,28,33,32,49]
flist1=[39]

#Define List
complete_list=[]
process_list=[]
complete_process_list=[]
error_list=[]
img_name=[]
img_source=[]
img_path=[]
img_date = []
img_width=[]
img_height=[]
img_channel=[]
img_size=[]
mask_name=[]
mask_path=[]
fill_mask=[]
fill_path=[]
hemo=[]
ma=[]
cws=[]
ex=[]
macula=[]
optic_disc=[]
optic_cup=[]
blood_vessel=[]
prh=[]
fvp=[]
micro_aneurysm=[]
drusen=[]
venous_beading=[]
laser_mark=[]
nve=[]
nvd=[]






#Column sheet name
original_data = {'Image Name': img_name,
'source': img_source,
'original_path':img_path,
'date':img_date,
'height':img_height,
'width':img_width,
'channel':img_channel,
'image bits': img_size}

#column data
data = {'mask_img':mask_name,
'mask_path':mask_path,
'mask_date':img_date,
'fill_mask_img':fill_mask,
'fill_mask_path':fill_path,
'HEMO': hemo,
'MA': ma,
'CWS': cws,
'EX': ex,
'macula':macula,
'optic disc':optic_disc,
'optic cup':optic_cup,
'blood vessel':blood_vessel,
'PRH':prh,
'FVP':fvp,
'micro aneurysms':micro_aneurysm,
'drusen':drusen,
'venous beading':venous_beading,
'laser marks':laser_mark,
'NVE':nve,
'NVD':nvd}


#Connection with IBM Bucket

def connect_cos():
    cos = ibm_boto3.client(service_name='s3',
                           ibm_api_key_id=cosCredentials['IBM_API_KEY_ID'],
                           ibm_service_instance_id=cosCredentials['IAM_SERVICE_ID'],
                           ibm_auth_endpoint=cosCredentials['IBM_AUTH_ENDPOINT'],
                           config=Config(signature_version='oauth'),
                           endpoint_url=cosCredentials['ENDPOINT'])
    return cos
    
# Upload the files from local directory with give file pattern to COS

def upload_files_cos(image_files_path,source_name,cate_type,local_dir,dt_object,local_error_dir,process_dir,process_dir_load,process_local_dir,pattern="*", bucket=defCOSBucketName,overwrite=False):
    #print("Uploading {0} files from {1} with overwrite {2}".format(pattern, local_dir, overwrite))
    file_list = []
    index = 0
    upload_count = 0

    for local_file in glob.glob(image_files_path + "/" + pattern):
        print("Processing {0}".format(local_file))

        index += 1
        filename = os.path.basename(local_file)
        print(filename)
        upload_file_cos(bucket, local_file,source_name,filename,image_files_path,dt_object,cate_type,process_dir)
        file_list.append(filename)
        upload_count +=1

    if len(mask_name)!=0:
    	cos = connect_cos()
    	for file in glob.glob(process_dir_load + "/" + pattern):
    		fill_file = os.path.basename(file)
    		image=cv2.imread(os.path.join(process_dir_load,fill_file))
    		image=cv2.cvtColor(image,cv2.COLOR_BGR2RGB)
    		imagef=image.copy()
    		gray = cv2.cvtColor(imagef, cv2.COLOR_RGB2GRAY)
    		mask2=generate_mask1(allfeatures,gray) #check missing boundary
    		cos.upload_file(Filename=file,Bucket=bucket,Key=source_name+"/fill_mask/"+fill_file)
    		fill_mask.append(fill_file)
    		fill_path.append(bucket+"/"+source_name+"/fill_mask/"+fill_file)
    		complete_process_list.append(fill_file)
    		fill_mask_feature(process_dir_load,fill_file)

    if len(complete_process_list)!=0:
    	dataframe_fun(match_count)


    print("Total files {0}/{1} uploaded to COS bucket {2}".format(len(complete_list), index, bucket))
    #print("local upload path:",image_files_path)
    complete_upload_file(image_files_path,local_dir)
    errorfiledir(image_files_path,local_error_dir)
    complete_upload_file1(process_dir_load,process_local_dir)
    
    return upload_count, file_list


# Upload individual file to COS

def upload_file_cos(bucket,local_file,source_name,filename,image_files_path,dt_object,cate_type,process_dir):
	cos = connect_cos()

	try:
		if cate_type == "mask":
			mask_auto(bucket,image_files_path,local_file,source_name,cate_type,filename,dt_object,process_dir)


		elif cate_type=="original":
			original_auto(bucket,image_files_path,local_file,source_name,cate_type,filename,dt_object)		
		


	except ConnectionError as e:
		error_list.append(filename)
		print(ConnectionError,e)

	except Exception as e:
		print(Exception, e)
	#else:
		#print('File {0} is uploaded to bucket {1}'.format(filename, bucket))

#MOVE COMPLETE FILE
def complete_upload_file1(image_files_path,local_dir):
	source=image_files_path
	destination=local_dir+"\\"
	#print("Total count:",len(complete_list))
	if len(complete_process_list)!=0:
		for item in complete_process_list:
			image_path=source+"\\"+item
			#print(image_path)
			if item in image_path:
				dest = shutil.move(image_path, destination+item)
	else:
		print("No file in complete_process_list")

#MOVE COMPLETE FILE
def complete_upload_file(image_files_path,local_dir):
	source=image_files_path
	destination=local_dir+"\\"
	if len(complete_list)!=0:
		print("Total success count:",len(complete_list))
		for item in complete_list:
			image_path=source+"\\"+item
			#print(image_path)
			if item in image_path:
				dest = shutil.move(image_path, destination+item)
	else:
		print("No file in complete_list")

# COPY ERROR FILE

def errorfiledir(image_files_path,local_error_dir):
	source=image_files_path
	destination=local_error_dir+"\\"
	if len(error_list)!=0:
		print("Total error count:",len(error_list))
		for item in error_list:
			image_path=source+"\\"+item
			#print(image_path)
			if item in image_path:
				dest = shutil.move(image_path, destination+item)
	else:
		print("No error file in directory")

#mask fill
def generate_mask(flist,gray,mask):
    for val in flist:
        if(int(val) in gray):
            h,w=gray.shape
            imghe=gray.copy()
            for i in range(h):
                for j in range(w):
                    if(imghe[i][j]==int(val)):
                        imghe[i][j]=255
                    else:
                        imghe[i][j]=0

            # find the contours from the thresholded image
            contours, hierarchy = cv2.findContours(imghe, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
            for cnt in contours:
                #print(cnt)
                x=[]
                y=[]
                color=(int(val),int(val),int(val))
                for item in cnt:
                    x.append(item[0][0])
                    y.append(item[0][1])
                xy_tuple=tuple(zip(x,y))
                polygon = [list(x) for x in xy_tuple]
                nppol=np.array([polygon], np.int32)
                pts=[nppol]
                cv2.fillPoly(mask,pts,color)


    return mask

def generate_mask1(allfeatures,gray):
	for val in allfeatures:
		if(int(val) in gray):
			h,w=gray.shape
			imghe=gray.copy()
			for i in range(h):
				for j in range(w):
					if(imghe[i][j]==int(val)):
						imghe[i][j]=255
					else:
						imghe[i][j]=0

			# find the contours from the thresholded image
			contours, hierarchy = cv2.findContours(imghe, cv2.RETR_CCOMP, cv2.CHAIN_APPROX_SIMPLE)
			color=(int(val),int(val),int(val))
			# Find the index of the largest contour
			areas = [cv2.contourArea(c) for c in contours]
			max_index = np.argmax(areas)
			cnt=contours[max_index]
			if hierarchy[0][max_index][3]==-1:
				print("this is not complete region")
				break

	return gray
    

def original_auto(bucket,image_files_path,local_file_name,source,cate_type,key,dt_object):
	cos = connect_cos()

	wb = openpyxl.load_workbook('image_registry.xlsx')		#read excel file
	sheet = wb['Sheet1']
	for rowNum in range(1, sheet.max_row+1):  # skip the first row
		update_col = sheet.cell(row=rowNum, column=1).value
		if key in update_col:
			print("match:",rowNum)
			print("File exists move ahead")
			print('File {0} has already uploaded in bucket {1}'.format(key, bucket))
			error_list.append(key)
			break
			
		else:
			if not key in update_col and (rowNum==sheet.max_row):
				cos.upload_file(Filename=local_file_name,Bucket=bucket,Key=source+"/"+cate_type+"/"+key)
				complete_list.append(key)
				img_name.append(key)
				img_source.append(source)
				img_date.append(dt_object)
				img_path.append(bucket+"/"+source+"/"+cate_type+"/"+key)
				if key in os.listdir(image_files_path):
					img=cv2.imread(os.path.join(image_files_path,key))
					if img is not None:									#append image height,width,channel and size in list
						img_height.append(img.shape[0])
						img_width.append(img.shape[1])
						img_channel.append(img.shape[2])
						img_size.append('{:,}'.format(img.size))
				rowNum=sheet.max_row+1
				index = original_data['Image Name'].index(key)
				print("{0} is append in sheet".format(key))
				print('File {0} is successfully uploaded to bucket {1}'.format(key, bucket))
				sheet.cell(row=rowNum, column = 1).value = original_data['Image Name'][index]
				sheet.cell(row=rowNum, column = 2).value = original_data['source'][index]
				sheet.cell(row=rowNum, column = 3).value = original_data['original_path'][index]
				sheet.cell(row=rowNum, column = 4).value = original_data['date'][index]
				sheet.cell(row=rowNum, column = 5).value = original_data['height'][index]
				sheet.cell(row=rowNum, column = 6).value = original_data['width'][index]
				sheet.cell(row=rowNum, column = 7).value = original_data['channel'][index]
				sheet.cell(row=rowNum, column = 8).value = original_data['image bits'][index]

	wb.save('image_registry.xlsx')

def mask_feature_extraction(mask_image_list):
	for file in mask_image_list:					#check features
		pixel_set=set(file.getdata())
		pixels={x for cordinate in pixel_set for x in cordinate}
		if 11 in pixels:				#check hemorrhages(11) 
			hemo.append('Y')
		else:
			hemo.append('N')
	
		if 15 in pixels:				#check ma(15) 
			ma.append('Y')
		else:
			ma.append('N')
	
		if 28 in pixels:				#check soft exudates(28) 
			cws.append('Y')
		else:
			cws.append('N')
	
		if 33 in pixels:				#check hard exudates(33) 
			ex.append('Y')
		else:
			ex.append('N')
	
		if 32 in pixels:				#check macula(32) 
			macula.append('Y')
		else:
			macula.append('N')
	
		if 49 in pixels:				#check optic disc(49) 
			optic_disc.append('Y')
		else:
			optic_disc.append('N')
	
		if 39 in pixels:				#check optic cup(39) 
			optic_cup.append('Y')
		else:
			optic_cup.append('N')

		if 72 in pixels:				#check blood vessels(72) 
			blood_vessel.append('Y')
		else:
			blood_vessel.append('N')

		if 29 in pixels:				#check pre retina haemorrhage(29) 
			prh.append('Y')
		else:
			prh.append('N')

		if 26 in pixels:				#check fibrovascular proliferations(26) 
			fvp.append('Y')
		else:
			fvp.append('N')

		if 15 in pixels:				#check micro aneurysms(15) 
			micro_aneurysm.append('Y')
		else:
			micro_aneurysm.append('N')

		if 14 in pixels:				#check drusen(14) 
			drusen.append('Y')
		else:
			drusen.append('N')

		if 13 in pixels:				#check venous beading(13) 
			venous_beading.append('Y')
		else:
			venous_beading.append('N')

		if 6 in pixels:				#check laser marks(6) 
			laser_mark.append('Y')
		else:
			laser_mark.append('N')

		if 5 in pixels:				#check neovascularization elsewhere(5) 
			nve.append('Y')
		else:
			nve.append('N')

		if 1 in pixels:				#check neovascularization of disc(1) 
			nvd.append('Y')
		else:
			nvd.append('N')


def mask_auto(bucket,image_files_path,local_file_name,source,cate_type,key,dt_object,process_dir):
	cos = connect_cos()

	wb = openpyxl.load_workbook('image_registry.xlsx')		#read excel file
	sheet = wb['Sheet1']
	for rowNum in range(2, sheet.max_row+1):  # skip the first row
		update_col = sheet.cell(row=rowNum, column=1).value
		mask_col = sheet.cell(row=rowNum, column=9).value
		if update_col[:-4] == key[:-9] and mask_col is None:
			print("match index:",rowNum)
			cos.upload_file(Filename=local_file_name,Bucket=bucket,Key=source+"/"+cate_type+"/"+key)
			complete_list.append(key)
			mask_name.append(key)
			mask_path.append(bucket+"/"+source+"/"+cate_type+"/"+key)
			img_date.append(dt_object)
			image=cv2.imread(os.path.join(image_files_path,key))
			image=cv2.cvtColor(image,cv2.COLOR_BGR2RGB)
			imagef=image.copy()
			gray = cv2.cvtColor(imagef, cv2.COLOR_RGB2GRAY)
			mask=np.zeros(image.shape, np.uint8)
			mask=generate_mask(flist,gray,mask)  #fill all features
			mask1=generate_mask(flist1,gray,mask) #fill only optic cup 
			cv2.imwrite(process_dir+key[:-4]+"_fill.png",mask1)	#file save in fill_mask
			break

			
		else:
			if not update_col[:-4] == key[:-9] and (rowNum==sheet.max_row):
				print("Mask file {0} has either no original file or mask image exist in sheet".format(key))
				error_list.append(key)
			else:
				if update_col[:-4] == key[:-9] and (rowNum==sheet.max_row):
					print("Mask file {0} match in last row of sheet".format(key))
					error_list.append(key)

				
	


def fill_mask_feature(process_dir_load,fill_file):
	path_list=listdir(process_dir_load)
	#mask image in list
	mask_image_list=[]
	if fill_file in os.listdir(process_dir_load):
		img = PImage.open(process_dir_load+"\\" + fill_file)
		#img=cv2.imread(os.path.join(process_dir_load,fill_file))
		if img is not None:
			mask_image_list.append(img)
	mask_feature_extraction(mask_image_list)


def dataframe_fun(match_count):
	df=pd.read_excel("image_registry.xlsx")
	#save path data into dataframe
	df_mask = pd.DataFrame(data)
	# insert_loc = df.index.max()
	# df_path=df.head(insert_loc)
	#modiDF1=df_path1.append(df_reg, ignore_index=True)
	print(df_mask)
	wb = openpyxl.load_workbook('image_registry.xlsx')
	sheet = wb['Sheet1']

	for rowNum in range(2, sheet.max_row+1):  # skip the first row
		update_col = sheet.cell(row=rowNum, column=1).value
		for i in range(len(df_mask)):
			key=df_mask['mask_img'][i]
			if update_col[:-4] == key[:-9]:
				match_count+=1
				index = i
				if sheet.cell(row=rowNum, column = 9).value is None:
					sheet.cell(row=rowNum, column = 9).value = data['mask_img'][index]
				if sheet.cell(row=rowNum, column = 10).value is None:
					sheet.cell(row=rowNum, column = 10).value = data['mask_path'][index]
				if sheet.cell(row=rowNum, column = 11).value is None:
					sheet.cell(row=rowNum, column = 11).value = data['mask_date'][index]
				if sheet.cell(row=rowNum, column = 12).value is None:
					sheet.cell(row=rowNum, column = 12).value = data['fill_mask_img'][index]
				if sheet.cell(row=rowNum, column = 13).value is None:
					sheet.cell(row=rowNum, column = 13).value = data['fill_mask_path'][index]
				if sheet.cell(row=rowNum, column = 14).value is None:
					sheet.cell(row=rowNum, column = 14).value = data['HEMO'][index]
				if sheet.cell(row=rowNum, column = 15).value is None:
					sheet.cell(row=rowNum, column = 15).value = data['MA'][index]
				if sheet.cell(row=rowNum, column = 16).value is None:
					sheet.cell(row=rowNum, column = 16).value = data['CWS'][index]
				if sheet.cell(row=rowNum, column = 17).value is None:
					sheet.cell(row=rowNum, column = 17).value = data['EX'][index]
				if sheet.cell(row=rowNum, column = 18).value is None:
					sheet.cell(row=rowNum, column = 18).value = data['macula'][index]
				if sheet.cell(row=rowNum, column = 19).value is None:
					sheet.cell(row=rowNum, column = 19).value = data['optic disc'][index]
				if sheet.cell(row=rowNum, column = 20).value is None:
					sheet.cell(row=rowNum, column = 20).value = data['optic cup'][index]
				if sheet.cell(row=rowNum, column = 21).value is None:
					sheet.cell(row=rowNum, column = 21).value = data['blood vessel'][index]
				if sheet.cell(row=rowNum, column = 22).value is None:
					sheet.cell(row=rowNum, column = 22).value = data['PRH'][index]
				if sheet.cell(row=rowNum, column = 23).value is None:
					sheet.cell(row=rowNum, column = 23).value = data['FVP'][index]
				if sheet.cell(row=rowNum, column = 24).value is None:
					sheet.cell(row=rowNum, column = 24).value = data['micro aneurysms'][index]
				if sheet.cell(row=rowNum, column = 25).value is None:
					sheet.cell(row=rowNum, column = 25).value = data['drusen'][index]
				if sheet.cell(row=rowNum, column = 26).value is None:
					sheet.cell(row=rowNum, column = 26).value = data['venous beading'][index]
				if sheet.cell(row=rowNum, column = 27).value is None:
					sheet.cell(row=rowNum, column = 27).value = data['laser marks'][index]
				if sheet.cell(row=rowNum, column = 28).value is None:
					sheet.cell(row=rowNum, column = 28).value = data['NVE'][index]
				if sheet.cell(row=rowNum, column = 29).value is None:
					sheet.cell(row=rowNum, column = 29).value = data['NVD'][index]
	

	wb.save('image_registry.xlsx')
	#print("Total match:",match_count)
if __name__ == "__main__":

	anno_img = connect_cos()

	source_name = input("Enter Image source name: ")
	cate_type = input("Enter Image category(original or mask): ")
	path="D:\\Aiseon\\Home\\AISeon_Image\\APTOS-VFI\\APTOS\\work\\source"
	com_path="D:\\Aiseon\\Home\\AISeon_Image\\APTOS-VFI\\APTOS\\work"

	# current date and time
	now = datetime.now()
	timestamp = datetime.timestamp(now)
	dt_object = datetime.fromtimestamp(timestamp)

	#Upload file path 
	image_files_path=path+"\\"+source_name+"\\"+cate_type

	#complete uploaded path
	local_dir=com_path+"\\"+"upload_completed"+"\\"+source_name+"\\"+cate_type
	process_local_dir=path+"\\"+source_name+"\\"+"fill_mask"
	#error uploaded path
	local_error_dir=com_path+"\\"+"upload_error"+"\\"+source_name+"\\"+cate_type
	#processed path
	process_dir=path+"\\"+source_name+"\\"+"fill_mask"+"\\"
	process_dir_load=path+"\\"+source_name+"\\"+"fill_mask"

	#path validation
	if not os.path.exists(image_files_path):
		print("Wrong file path")
	else:
		#Upload image file in bucket
		upload_files_cos(image_files_path,source_name,cate_type,local_dir,dt_object,local_error_dir,process_dir,process_dir_load,process_local_dir)
